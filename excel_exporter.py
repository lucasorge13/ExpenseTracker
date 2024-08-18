import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import PieChart, Reference
import csv
import datetime

def exportExpensesToExcel(filePath, budget):
    print("Exporting expenses to Excel...")
    excel_file_path = 'expense_summary.xlsx'
    
    # Create a new Excel workbook and add a worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Monthly Expenses"

    # Add headers to the sheet
    headers = ["Date", "Category", "Expense Name", "Amount (USD)"]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # Dictionary to store expenses by date and category
    expenses_by_date = {}
    category_totals = {}

    # Read data from the expense CSV file
    with open(filePath, "r") as file:
        reader = csv.DictReader(file)
        for row in reader:
            try:
                name = row["Expense Name"]
                category = row["Category"]
                amount = float(row["Amount"])
                date = datetime.datetime.now().strftime("%d %B %Y")

                # Store expenses by date and category
                if date not in expenses_by_date:
                    expenses_by_date[date] = {}
                if category not in expenses_by_date[date]:
                    expenses_by_date[date][category] = []

                expenses_by_date[date][category].append({"name": name, "amount": amount})

                # Calculate total per category
                if category not in category_totals:
                    category_totals[category] = 0
                category_totals[category] += amount
            except ValueError as e:
                print(f"Skipping invalid row: {row}, Error: {e}")

    # Write data to the Excel sheet
    row_num = 2
    total_spent = 0
    border_style = Border(left=Side(style='thin'), 
                          right=Side(style='thin'), 
                          top=Side(style='thin'), 
                          bottom=Side(style='thin'))

    for date_str, categories in expenses_by_date.items():
        for category, expenses in categories.items():
            for expense in expenses:
                sheet.cell(row=row_num, column=1, value=date_str).border = border_style
                sheet.cell(row=row_num, column=2, value=category).border = border_style
                sheet.cell(row=row_num, column=3, value=expense["name"]).border = border_style
                sheet.cell(row=row_num, column=4, value=expense["amount"]).border = border_style
                total_spent += expense["amount"]
                row_num += 1

    # Add summary information with formatting
    summary_start_row = row_num + 2

    summary_cells = [
        {"label": "Total Budget:", "value": budget, "color": "000000"},  # Black text for the budget
        {"label": "Total Spent:", "value": total_spent, "color": "1F4E78"},  # Blue text for the total spent
        {"label": "Remaining Budget:", "value": budget - total_spent, "color": "FF0000"}  # Red text for remaining budget
    ]

    for i, summary in enumerate(summary_cells):
        label_cell = sheet.cell(row=summary_start_row + i, column=3, value=summary["label"])
        label_cell.font = Font(bold=True)
        label_cell.alignment = Alignment(horizontal="right")
        label_cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        label_cell.border = border_style

        value_cell = sheet.cell(row=summary_start_row + i, column=4, value=summary["value"])
        value_cell.font = Font(bold=True, color=summary["color"])
        value_cell.alignment = Alignment(horizontal="center")
        value_cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        value_cell.border = border_style

    # Insert a pie chart for category expenses
    chart_start_row = summary_start_row + 5
    category_row_start = chart_start_row

    for i, (category, total) in enumerate(category_totals.items(), start=category_row_start):
        sheet.cell(row=i, column=2, value=category).font = Font(bold=True)
        sheet.cell(row=i, column=4, value=total)
    
    # Add remaining budget to the pie chart data
    remaining_budget = budget - total_spent
    sheet.cell(row=i + 1, column=2, value="Remaining Budget").font = Font(bold=True)
    sheet.cell(row=i + 1, column=4, value=remaining_budget)
    
    pie = PieChart()
    labels = Reference(sheet, min_col=2, min_row=category_row_start, max_row=i + 1)
    data = Reference(sheet, min_col=4, min_row=category_row_start, max_row=i + 1)
    pie.add_data(data, titles_from_data=False)
    pie.set_categories(labels)
    pie.title = "Expenses by Category and Budget"
    pie.height = 8  # Chart height in inches
    pie.width = 10   # Chart width in inches

    # Apply a custom color palette using the 'style' attribute to simulate different colors
    pie.style = 10  # Uses one of openpyxl's built-in styles with varied colors

    # Move the legend to the right side with some spacing
    pie.legend.position = 'r'  # Moves the legend to the right
    pie.legend.overlay = False  # Ensures the legend doesn't overlay the chart
    pie.legend.spacing = 1.5  # Adjusts the spacing between the legend and the chart

    sheet.add_chart(pie, f"E{chart_start_row}")

    # Enhance the table next to the pie chart
    for k in range(category_row_start, i + 2):
        for col in range(2, 5):  # Adjusting Category, Total, and Budget cells
            cell = sheet.cell(row=k, column=col)
            cell.border = border_style
            if col == 2:
                cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

    # Adjust column width for better readability
    for col in sheet.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        sheet.column_dimensions[col[0].column_letter].width = max_length + 2

    # Save the workbook
    workbook.save(excel_file_path)
    print(f"Expenses have been exported to {excel_file_path}")
